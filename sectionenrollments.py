import requests
import json
import time
import openpyxl


course_id = '537'  
section_id = '1433'
file_name = 'Z9.xlsx'  

def start_execution():
    print(f'Course ID : {course_id}')
    time.sleep(1.25)
    print(f'Section ID : {section_id}')
    time.sleep(1.25)
    print(f'Batch File Name : {file_name}')
    time.sleep(1.55)

CANVAS_API_URL = "https://canvas.true-education.org/api/v1"
CANVAS_API_TOKEN = "16479~zu8TB4WvxVcry6TcLhvJYUWnRrMN6uTLPnxcRXACYW8ZKNtWem4M9EFvTJAaPD6k"

def get_student_ids(file_name):
    dataframe = openpyxl.load_workbook(file_name)
    dataframe1 = dataframe.active

    students = []
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            students.append(col[row].value)

    for student in students:
     time.sleep(0.25)
     print(student)
    
    return students

def enroll_user_in_section(course_id, section_id, user_id, enrollment_type='StudentEnrollment'):
    url = f'{CANVAS_API_URL}/sections/{section_id}/enrollments'
    headers = {
        'Authorization': f'Bearer {CANVAS_API_TOKEN}',
        'Content-Type': 'application/json'
    }
    payload = {
        'enrollment': {
            'user_id': user_id,
            'type': enrollment_type,
            'enrollment_state': 'active'
        }
    }

    response = requests.post(url, headers=headers, json=payload)
    return response.json()

def unenroll_user(course_id, enrollment_id):
    url = f"{CANVAS_API_URL}/courses/{course_id}/enrollments/{enrollment_id}"
    response = requests.delete(url, headers=get_headers())
    response.raise_for_status()
    print("User successfully unenrolled.")

def enroll_students():
    student_ids = get_student_ids(file_name)
    for student_id in student_ids:
        time.sleep(0.25)
        response = enroll_user_in_section(course_id, section_id, student_id)
        print(response)

start_execution()
enroll_students()
